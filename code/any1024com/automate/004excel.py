# coding=utf-8
import pathlib
import random
import datetime
import calendar
from openpyxl import Workbook, load_workbook
from openpyxl.styles import numbers
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList


def auto_chart(in_path, out_path):
    """
    自动生成报告
    汇总某个文件夹内所有同类Excel文件，输出报表
    以时间维度，粒度为月，汇总统计单店经营和全国经营
    输出报表：
        1. 各门店经营：销售额、成本、毛利
        2. 全国经营：销售额，成本，毛利
        3. 门店总排行榜：销售额、毛利润
    TODO: 可以增加产品维度的统计作为练习。

    Version: 0.0.1
    Author : yichu.cheng
    """
    wb = Workbook()
    ws_all = wb.active
    ws_all.title = '全国汇总'
    # 一、先为每个门店生成一张单店经营表
    # 月份, 总营收, 总折扣, 总成本, 总毛利
    shop_list = list(in_path.glob('*.xlsx'))
    shop_names = [shop.stem for shop in shop_list]
    for shop in shop_list:
        # shop = shop_list[0]
        print(f'正在处理 {shop.stem} 表......')
        wb_shop = load_workbook(shop)  # 打开城市门店数据
        ws_shop = wb_shop.active
        shop_name = shop.stem
        ws = wb.create_sheet(shop_name)  # 生成对应城市门店表
        # 表头
        header_list = ['月份', '销售额', '折扣额', '成本', '毛利润']
        ws.append(header_list)
        # 从1到12月生成数据汇总表
        for m in range(1, 13):
            start_date = datetime.datetime(2020, m, 1)
            _, days_in_month = calendar.monthrange(
                start_date.year, start_date.month)
            end_date = start_date + datetime.timedelta(days=days_in_month)
            # print(m, start_date, end_date)
            # 按月提取数据
            month_data_list = [
                ([ws_shop.cell(row=i, column=j).value for j in range(6, 10)])
                for i in range(2, ws_shop.max_row + 1)
                if start_date <= ws_shop.cell(row=i, column=1).value < end_date
            ]
            # 累加汇总
            # zip(*list)可以把list中的tuple重新按序打包
            month_data_T = list(zip(*month_data_list))
            month_sum = [
                f'{m}月',
                sum(month_data_T[1]),  # sum of sale
                # sum of sale*(1-discount) 折扣总额
                sum([a*(1-b) for a, b in zip(month_data_T[1], month_data_T[0])]),
                sum(month_data_T[2]),  # sum of cost
                sum(month_data_T[3])  # sum of profit
            ]
            ws.append(month_sum)
        year_sum = [
            '全年总计',
            '=SUM(B2:B13)',
            '=SUM(C2:C13)',
            '=SUM(D2:D13)',
            '=SUM(E2:E13)'
        ]
        ws.append(year_sum)
        # 画个月度统计图表
        # 列状图
        bar_chart = BarChart()
        bar_chart.type = 'col'  # 列状图
        bar_chart.style = 10
        bar_chart.title = '门店经营统计图'
        bar_chart.y_axis.title = '金额'
        bar_chart.x_axis.title = '月份'
        d_ref = Reference(ws, min_col=2, min_row=1, max_row=13, max_col=5)
        cat = Reference(ws, min_col=1, min_row=2, max_row=13)
        bar_chart.add_data(d_ref, titles_from_data=True)
        bar_chart.set_categories(cat)
        ws.add_chart(bar_chart, 'A17')

    # 开始处理全国总表
    print('开始处理全国总表......')
    ws_all.append(header_list)
    for m in range(1, 13):
        # (销售额, 折扣额, 成本, 毛利润)
        month_all_shop = [
            ([wb[shop].cell(row=m+1, column=j).value
              for j in range(2, 6)])
            for shop in shop_names
        ]
        month_all_shop_T = list(zip(*month_all_shop))
        month_all_sum = [
            f'{m}月',
            sum(month_all_shop_T[0]),  # sale
            sum(month_all_shop_T[1]),  # discount
            sum(month_all_shop_T[2]),  # cost
            sum(month_all_shop_T[3]),  # profit
        ]
        ws_all.append(month_all_sum)
    ws_all.append(year_sum)

    # 生成全国月度统计报表
    bar_chart = BarChart()
    bar_chart.type = 'col'  # 列状图
    bar_chart.style = 10
    bar_chart.title = '全国经营统计图'
    bar_chart.y_axis.title = '金额'
    bar_chart.x_axis.title = '月份'
    d_ref = Reference(ws_all, min_col=2, min_row=1, max_row=13, max_col=5)
    cat = Reference(ws_all, min_col=1, min_row=2, max_row=13)
    bar_chart.add_data(d_ref, titles_from_data=True)
    bar_chart.set_categories(cat)
    ws_all.add_chart(bar_chart, 'J1')

    # 生成城市门店排名数据
    header_list = ['城市门店', '年度销售额', '年度毛利润']
    for j in range(3):
        ws_all.cell(row=19, column=j+1, value=header_list[j])
    for i, shop in enumerate(shop_names, start=2):
        ws_shop = wb[shop]
        year_sale = sum([ws_shop.cell(row=r, column=2).value
                         for r in range(2, 14)])
        year_profit = sum([ws_shop.cell(row=r, column=5).value
                           for r in range(2, 14)])
        ws_all.append([
            shop, year_sale, year_profit
        ])

    # 排序
    ws_all.auto_filter.ref = 'A19:C27'
    ws_all.auto_filter.add_sort_condition('B20:B27')
    # 生成排名图表
    bar_chart = BarChart()
    bar_chart.type = 'col'
    bar_chart.style = 10
    bar_chart.title = '城市门店年度经营'
    bar_chart.y_axis.title = '年度销售额'
    bar_chart.x_axis.title = '城市门店'
    d_ref = Reference(ws_all, min_col=2, min_row=19, max_row=27, max_col=3)
    cat = Reference(ws_all, min_col=1, min_row=20, max_row=27)
    bar_chart.add_data(d_ref, titles_from_data=True)
    bar_chart.set_categories(cat)
    ws_all.add_chart(bar_chart, 'J19')
    # 门店销售额比例饼图
    pie_chart = PieChart()
    pie_chart.dataLabels = DataLabelList()
    pie_chart.dataLabels.showPercent = True
    pie_chart.style = 10
    pie_chart.title = '门店销量占比'
    d_ref = Reference(ws_all, min_col=2, min_row=19, max_row=27)
    cat = Reference(ws_all, min_col=1, min_row=20, max_row=27)
    pie_chart.add_data(d_ref, titles_from_data=True)
    pie_chart.set_categories(cat)
    ws_all.add_chart(pie_chart, 'S1')
    # 门店利润比例饼图
    pie_chart = PieChart()
    pie_chart.dataLabels = DataLabelList()
    pie_chart.dataLabels.showPercent = True
    pie_chart.title = '门店利润占比'
    d_ref = Reference(ws_all, min_col=3, min_row=19, max_row=27)
    cat = Reference(ws_all, min_col=1, min_row=20, max_row=27)
    pie_chart.add_data(d_ref, titles_from_data=True)
    pie_chart.set_categories(cat)
    ws_all.add_chart(pie_chart, 'S19')

    wb.save(out_path)


def gen_data(out_path):
    """
    假设我们正在经营一个连锁茶饮品牌，也叫`Python1024`
    生成样本数据，模型源于实践应用，数据随机生成
    生成多个城市门店经营数据，粒度到日
    包括产品、订单、单价、销售额、折扣、产品成本等维度

    Version: 0.0.1
    Author : yichu.cheng
    """
    shop_list = ['广州', '深圳', '厦门', '上海', '杭州', '北京', '武汉', '湖南']
    product_list = ['柠檬蜂蜜绿茶', '波霸奶茶', '芒果冰沙', '布丁奶茶', '牛油草莓派']
    price_list = [19, 22, 24, 19, 30]
    cost_list = [7, 5, 8, 7, 12]
    discount_list = [0.5, 0.7, 0.8, 0.9, 1]

    for shop in shop_list:
        print(f'正在生成 {shop} 店铺数据......')
        wb = Workbook()
        ws = wb.active
        head_list = ['2020年', '订单ID', '产品', '单价',
                     '数量', '折扣', '销售额', '成本', '毛利', '毛利率']
        ws.append(head_list)
        ws.freeze_panes = ws['A1']
        start_date = datetime.date(2020, 1, 1)
        delta_date = datetime.timedelta(days=1)
        row_index = 1
        for i in range(366):
            date = start_date + delta_date * i
            # 每天有[5, 30]个订单
            cnt_order = random.randint(5, 30)
            for o in range(cnt_order):
                order_id = f'{date}-{o}'
                product = random.choice(product_list)
                p_index = product_list.index(product)
                price = price_list[p_index]
                discount = random.choice(discount_list)
                num = random.randint(1, 10)  # 1-10杯
                sale = price * num
                cost = cost_list[p_index] * num
                profit = sale * discount - cost
                profit_rate = profit * 1.0 / sale

                order = [date, order_id, product, price, num,
                         discount, sale, cost, profit, profit_rate]
                ws.append(order)
                # 简单设置数据格式
                row_index += 1
                ws.cell(row=row_index,
                        column=1).number_format = numbers.FORMAT_DATE_YYYYMMDD2
                ws.cell(row=row_index,
                        column=6).number_format = numbers.FORMAT_PERCENTAGE
                ws.cell(row=row_index,
                        column=10).number_format = numbers.FORMAT_PERCENTAGE

        wb.save(out_path.joinpath(f'{shop}.xlsx'))


if __name__ == "__main__":
    path = pathlib.Path().cwd().joinpath('data/automate/004excel')
    in_folder = path.joinpath('004excel_case_in')
    out_path = path.joinpath('004excel_case_out.xlsx')
    if not in_folder.is_dir():
        in_folder.mkdir()
    # gen_data(in_folder)
    auto_chart(in_folder, out_path)
